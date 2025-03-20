import random
import io
import sys
import pandas as pd
import os
from datetime import datetime
from loguru import logger
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from stake_calculator import calculate_stakes_with_strategy, StakePair
import copy

# Define number of simulations to run
NUM_SIMULATIONS = 10

# Setup custom log handler to capture logs
class LogCapture:
    def __init__(self):
        self.logs = io.StringIO()
        self.handler_id = None

    def start(self):
        self.logs = io.StringIO()
        self.handler_id = logger.add(self.logs, level="INFO", filter=lambda record: record["name"] == "stake_calculator")

    def stop(self):
        if self.handler_id:
            logger.remove(self.handler_id)
            self.handler_id = None

    def get_logs(self):
        return self.logs.getvalue()


def read_from_excel():
    """
    Try to read player data from an examplebets.xlsx file in the current directory.
    Returns a tuple of (players_dict, cap_info_dict) if successful, or (None, None) if not.
    """
    filename = "examplebets.xlsx"
    
    if not os.path.exists(filename):
        print(f"File '{filename}' not found in current directory.")
        return None, None
    
    try:
        # Read the Excel file
        df = pd.read_excel(filename)
        
        # Check if required columns exist
        required_columns = ["Player Name", "Player Bet"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            print(f"Missing required columns in '{filename}': {', '.join(missing_columns)}")
            return None, None
        
        # Create players dictionary
        players = {}
        cap_info = {}
        
        for _, row in df.iterrows():
            player_name = str(row["Player Name"])
            player_bet = int(row["Player Bet"])
            
            # Check if cap setting column exists
            if "Cap Player Settings" in df.columns:
                # Handle different possible values in the Cap Player Settings column
                cap_setting = row["Cap Player Settings"]
                is_capped = True  # Default to True
                
                if isinstance(cap_setting, str):
                    cap_setting = cap_setting.lower()
                    if cap_setting in ['no', 'false', 'off', '0']:
                        is_capped = False
                elif isinstance(cap_setting, bool):
                    is_capped = cap_setting
                elif isinstance(cap_setting, (int, float)):
                    is_capped = bool(cap_setting)
                
                cap_info[player_name] = is_capped
            else:
                # Default to True if column doesn't exist
                cap_info[player_name] = True
            
            players[player_name] = player_bet
        
        print(f"Successfully loaded {len(players)} players from '{filename}'")
        # Print cap settings for reference
        for player, is_capped in cap_info.items():
            cap_status = "CAPPED" if is_capped else "UNCAPPED"
            print(f"  • {player}: {players[player]} tix - {cap_status}")
        
        return players, cap_info
    
    except Exception as e:
        print(f"Error reading from Excel file '{filename}': {str(e)}")
        return None, None


def get_player_stakes():
    """
    Get user input for player stakes.
    Returns a tuple of (players_dict, cap_info_dict).
    """
    print("Enter player stakes:")
    
    players = {}
    cap_info = {}
    player_count = 0
    
    while True:
        player_input = input(f"Player {player_count + 1} name, stake, and cap setting (e.g., 'Alice 100 yes') or 'done' to finish: ")
        
        if player_input.lower() == 'done':
            break
            
        try:
            parts = player_input.split(' ')
            if len(parts) < 2:
                print("Invalid format. Please enter at least name and stake separated by space.")
                continue
                
            name = parts[0]
            stake = int(parts[1])
            
            # Check for cap setting (default to True if not provided)
            is_capped = True
            if len(parts) > 2:
                cap_setting = parts[2].lower()
                if cap_setting in ['no', 'false', 'off', '0', 'n']:
                    is_capped = False
            
            # Use name as player ID
            player_id = name
            players[player_id] = stake
            cap_info[player_id] = is_capped
            player_count += 1
            
            # Print confirmation with cap status
            cap_status = "CAPPED" if is_capped else "UNCAPPED"
            print(f"Added {name}: {stake} tix - {cap_status}")
            
            # For testing, we'll allow 6 or 8 players
            if player_count >= 8:
                break
        except ValueError:
            print("Invalid stake amount. Please enter a number.")
    
    # Validate we have enough players
    if player_count < 6:
        print("Need at least 6 players to run simulations.")
        return None, None
    
    # If odd number of players between 6 and 8, ask for one more
    if player_count % 2 != 0:
        print("Need an even number of players. Please add one more.")
        return get_player_stakes()
        
    return players, cap_info


def randomize_teams(player_ids):
    """
    Randomly split players into two equal teams.
    """
    # Convert to list and shuffle
    players = list(player_ids)
    random.shuffle(players)
    
    # Split into two equal teams
    mid_point = len(players) // 2
    team_a = players[:mid_point]
    team_b = players[mid_point:]
    
    return team_a, team_b


def run_stake_simulations(players, cap_info, min_stake=10, multiple=10):
    """
    Run multiple stake calculation simulations with different team assignments.
    Returns simulation results and captured logs.
    """
    results = []
    all_logs = []
    player_ids = list(players.keys())
    
    log_capture = LogCapture()
    
    for i in range(NUM_SIMULATIONS):
        print(f"Running simulation {i+1}...")
        
        # Randomize teams
        team_a, team_b = randomize_teams(player_ids)
        
        # Start log capture
        log_capture.start()
        
        # Create a deep copy of the stakes dictionary to prevent modifications from affecting future runs
        stakes_copy = copy.deepcopy(players)
        
        # Run stake calculation using tiered algorithm with cap_info
        tiered_pairs = calculate_stakes_with_strategy(
            team_a=team_a,
            team_b=team_b,
            stakes=stakes_copy,  # Use the copy instead of the original
            min_stake=min_stake,
            multiple=multiple,
            cap_info=cap_info  # Pass the cap_info parameter
        )
        
        # Stop log capture
        log_capture.stop()
        sim_log = log_capture.get_logs()
        
        # Calculate total bets per player
        player_bets = {}
        
        # Calculate totals
        for pair in tiered_pairs:
            player_bets[pair.player_a_id] = player_bets.get(pair.player_a_id, 0) + pair.amount
            player_bets[pair.player_b_id] = player_bets.get(pair.player_b_id, 0) + pair.amount
        
        # Format player bet summaries
        bet_summary = []
        
        for player_id in sorted(players.keys()):
            max_stake = players[player_id]  # Use original max stake for percentage calculation
            bet = player_bets.get(player_id, 0)
            
            # Calculate bet percentages
            pct = (bet / max_stake) * 100 if max_stake > 0 else 0
            
            # Add cap status to the summary
            cap_status = "🧢" if cap_info.get(player_id, True) else "🏎️"
            bet_summary.append(f"{player_id} {cap_status}: {bet}/{max_stake} ({pct:.1f}%)")
        
        # Store results
        sim_result = {
            'simulation': i+1,
            'team_a': [f"{p} ({players[p]} tix)" + (" 🧢" if cap_info.get(p, True) else " 🏎️") for p in team_a],
            'team_b': [f"{p} ({players[p]} tix)" + (" 🧢" if cap_info.get(p, True) else " 🏎️") for p in team_b],
            'stake_pairs': [f"{p.player_a_id} vs {p.player_b_id}: {p.amount} tix" for p in tiered_pairs],
            'total_stake': sum(p.amount for p in tiered_pairs),
            'num_pairs': len(tiered_pairs),
            'player_bets': player_bets,
            'bet_summary': bet_summary,
            'modified_stakes': stakes_copy,  # Store the modified stakes for reference
            'cap_info': cap_info  # Store cap info for reference
        }
        
        results.append(sim_result)
        all_logs.append(sim_log)
    
    return results, all_logs


def write_to_excel(players, cap_info, results, logs):
    """
    Write simulation results to Excel file.
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Add input sheet
    input_sheet = wb.create_sheet("Player Inputs")
    input_sheet.append(["Player", "Max Stake", "Cap Setting"])
    
    for player, stake in sorted(players.items(), key=lambda x: x[1], reverse=True):
        cap_setting = "CAPPED" if cap_info.get(player, True) else "UNCAPPED"
        input_sheet.append([player, stake, cap_setting])
    
    # Format headers
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in input_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add summary sheet
    summary_sheet = wb.create_sheet("Summary", 0)  # Make it the first sheet
    summary_sheet.append([
        "Simulation", 
        "Total Stake", 
        "# Transactions",
        "Team A", 
        "Team B"
    ])
    
    # Format header
    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add results for each simulation
    for i, (result, log) in enumerate(zip(results, logs)):
        # Create simulation sheet
        sim_sheet = wb.create_sheet(f"Sim {i+1}")
        
        # Add team information
        sim_sheet.append(["Team A (sorted by stake)"])
        # Sort team A by stake amount
        team_a_with_stakes = [(p.split(" (")[0], int(p.split("(")[1].split(" ")[0])) for p in result['team_a']]
        team_a_with_stakes.sort(key=lambda x: x[1], reverse=True)
        team_a_sorted = result['team_a']  # Keep original format with cap status
        for player in team_a_sorted:
            sim_sheet.append([player])
        
        sim_sheet.append([])  # Empty row
        
        sim_sheet.append(["Team B (sorted by stake)"])
        # Sort team B by stake amount
        team_b_with_stakes = [(p.split(" (")[0], int(p.split("(")[1].split(" ")[0])) for p in result['team_b']]
        team_b_with_stakes.sort(key=lambda x: x[1], reverse=True)
        team_b_sorted = result['team_b']  # Keep original format with cap status
        for player in team_b_sorted:
            sim_sheet.append([player])
        
        sim_sheet.append([])  # Empty row
        
        # Add stake pairs
        sim_sheet.append(["Tiered Stake Pairs"])
        for pair in result['stake_pairs']:
            sim_sheet.append([pair])
        
        sim_sheet.append([])  # Empty row
        sim_sheet.append(["Total Stake", result['total_stake']])
        sim_sheet.append(["Number of Transactions", result['num_pairs']])
        
        # Add player bet details
        sim_sheet.append([])  # Empty row
        sim_sheet.append(["Player Bet Details"])
        
        # Format as a table
        sim_sheet.append(["Player", "Cap Setting", "Max Stake", "MTMB Adjusted Stake", "Allocated", "Percentage"])
        
        # Get modified stakes from this simulation
        modified_stakes = result.get('modified_stakes', {})
        
        for player_id in sorted(players.keys()):
            original_max_stake = players[player_id]
            adjusted_stake = modified_stakes.get(player_id, original_max_stake)
            allocated = result['player_bets'].get(player_id, 0)
            percentage = (allocated / original_max_stake * 100) if original_max_stake > 0 else 0
            cap_setting = "CAPPED" if cap_info.get(player_id, True) else "UNCAPPED"
            
            # Only show adjusted stake if it differs from original
            if original_max_stake != adjusted_stake:
                sim_sheet.append([player_id, cap_setting, original_max_stake, adjusted_stake, allocated, f"{percentage:.1f}%"])
            else:
                sim_sheet.append([player_id, cap_setting, original_max_stake, "-", allocated, f"{percentage:.1f}%"])
        
        # Format headers
        for row_idx in [1, 4, 8, len(team_a_sorted) + len(team_b_sorted) + 8]:
            try:
                for cell in sim_sheet[row_idx]:
                    cell.fill = header_fill
                    cell.font = header_font
            except IndexError:
                # Skip if row doesn't exist
                pass
        
        # Add MTMB section to show what happened with stakes
        sim_sheet.append([])  # Empty row
        sim_sheet.append(["MTMB and Cap Adjustments"])
        sim_sheet.append(["Player", "Original Stake", "Cap Setting", "MTMB Adjusted"])
        
        for player_id in sorted(modified_stakes.keys()):
            original = players[player_id]
            adjusted = modified_stakes[player_id]
            cap_setting = "CAPPED" if cap_info.get(player_id, True) else "UNCAPPED"
            if original != adjusted:
                sim_sheet.append([player_id, original, cap_setting, adjusted])
        
        # Add logs
        log_sheet = wb.create_sheet(f"Sim {i+1} Logs")
        
        # Process logs
        log_lines = log.strip().split('\n')
        for line in log_lines:
            log_sheet.append([line])
        
        # Format header
        log_sheet.cell(row=1, column=1).fill = header_fill
        log_sheet.cell(row=1, column=1).font = header_font
        
        # Add to summary - shorten team listings
        summary_sheet.append([
            result['simulation'],
            result['total_stake'],
            result['num_pairs'],
            ", ".join([p.split(" (")[0] for p in team_a_sorted]),
            ", ".join([p.split(" (")[0] for p in team_b_sorted])
        ])
    
    # Adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 80 else 80
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tiered_stake_simulations_{timestamp}.xlsx"
    
    # Save workbook
    wb.save(filename)
    print(f"Results saved to {filename}")
    return filename


def main():
    print("Welcome to the Tiered Stake Algorithm Tester")
    print("===========================================")
    
    # First try to read from Excel file
    players, cap_info = read_from_excel()
    
    # If Excel file not found or invalid, use manual input
    if not players:
        print("\nEntering manual input mode...")
        players, cap_info = get_player_stakes()
        if not players:
            return
    
    # Get min stake
    try:
        min_stake = int(input("Enter minimum stake (default 10): ") or "10")
    except ValueError:
        min_stake = 10
        print("Invalid input. Using default minimum stake of 10.")
    
    # Get multiple
    try:
        multiple = int(input("Enter rounding multiple (default 10): ") or "10")
    except ValueError:
        multiple = 10
        print("Invalid input. Using default multiple of 10.")
    
    # Run simulations with the tiered algorithm
    results, logs = run_stake_simulations(players, cap_info, min_stake, multiple)
    
    # Write to Excel
    filename = write_to_excel(players, cap_info, results, logs)
    
    print(f"Testing complete! Results saved to {filename}")


if __name__ == "__main__":
    main()